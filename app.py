#!/usr/bin/env python3
"""
NmapViz - Graphical nmap results visualizer (BloodHound-style)
Port: 12221
"""

import xml.etree.ElementTree as ET
import json
import os
import re
import io
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, Response

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

HISTORY_DIR = Path('history')
HISTORY_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# PORT CLASSIFICATION  (critical = direct vulnerability just by being exposed)
# ─────────────────────────────────────────────────────────────────────────────

CRITICAL_PORTS = {
    # ── Cleartext remote access ───────────────────────────────────────────
    23:    {"name": "Telnet",        "reason": "No encryption. Credentials sent in plaintext. Replace with SSH."},
    512:   {"name": "rexec",         "reason": "Remote execution without encryption. Obsolete and dangerous."},
    513:   {"name": "rlogin",        "reason": "Remote login without encryption. Replaced by SSH."},
    514:   {"name": "rsh/syslog",    "reason": "Remote shell with no authentication. Extremely dangerous."},
    # ── Cleartext file transfer ───────────────────────────────────────────
    21:    {"name": "FTP",           "reason": "File transfer with no encryption. Credentials exposed."},
    69:    {"name": "TFTP",          "reason": "No authentication. Allows arbitrary file read/write."},
    873:   {"name": "rsync",         "reason": "Can allow unauthenticated file access depending on config."},
    # ── Legacy info disclosure ────────────────────────────────────────────
    79:    {"name": "Finger",        "reason": "Reveals system users. Facilitates enumeration attacks."},
    # ── SNMP / network management ─────────────────────────────────────────
    161:   {"name": "SNMP UDP",      "reason": "Default community strings (public/private). Full device info exposed."},
    162:   {"name": "SNMP Trap",     "reason": "SNMP trap receiver. Sensitive network events may be intercepted."},
    # ── Windows / SMB legacy ─────────────────────────────────────────────
    135:   {"name": "MSRPC",         "reason": "Windows RPC. History of critical CVEs (MS03-026, etc)."},
    137:   {"name": "NetBIOS-NS",    "reason": "Enables LLMNR/NBT-NS poisoning attacks."},
    138:   {"name": "NetBIOS-DGM",   "reason": "NetBIOS datagrams. Network information exposed."},
    # ── NFS / RPC ─────────────────────────────────────────────────────────
    111:   {"name": "RPCbind",       "reason": "Allows RPC service enumeration. Frequent attack vector."},
    2049:  {"name": "NFS",           "reason": "File sharing potentially without proper authentication."},
    # ── ICS / SCADA protocols (should NEVER be internet-exposed) ─────────
    102:   {"name": "S7comm (Siemens)","reason": "Siemens PLC protocol. Direct ICS/OT control possible if reachable."},
    502:   {"name": "Modbus",        "reason": "ICS protocol with no authentication. Direct hardware control."},
    20000: {"name": "DNP3",          "reason": "ICS/SCADA protocol. No native authentication."},
    44818: {"name": "EtherNet/IP",   "reason": "Industrial Ethernet protocol. CIP direct device access."},
    47808: {"name": "BACnet",        "reason": "Building automation protocol. Device control possible."},
    4840:  {"name": "OPC-UA",        "reason": "Industrial OPC protocol. Verify authentication is enforced."},
    # ── Network device management ─────────────────────────────────────────
    623:   {"name": "IPMI",          "reason": "BMC/IPMI. Authentication bypass CVEs (CVE-2013-4786). Remote OS control."},
    4786:  {"name": "Cisco Smart Install","reason": "Unauthenticated. Allows complete device takeover (IOS configs, RCE)."},
    7911:  {"name": "OMAPI (DHCP)",  "reason": "OMAPI interface for ISC DHCP. Can modify DHCP leases without auth."},
    9100:  {"name": "JetDirect / RAW Print","reason": "HP JetDirect raw printing. Can read/write printer file system."},
    # ── Backdoor / common malware ports ──────────────────────────────────
    4444:  {"name": "Backdoor",      "reason": "Commonly used by reverse shells and backdoors (Metasploit default)."},
    5554:  {"name": "Backdoor",      "reason": "Associated with backdoors and historical worms."},
    6666:  {"name": "IRC/Backdoor",  "reason": "Frequently used by botnets and IRC-based malware."},
    6667:  {"name": "IRC",           "reason": "Unencrypted IRC. Frequently exploited by botnets."},
    1524:  {"name": "Ingreslock Backdoor","reason": "Classic backdoor port. No legitimate services should use this."},
    # ── X11 (display servers) ─────────────────────────────────────────────
    6000:  {"name": "X11",           "reason": "X Window System. Can allow screen capture and input injection."},
    6001:  {"name": "X11 (display 1)","reason": "X Window System alternate display. Same risks as 6000."},
    # ── Compiler / debug services ─────────────────────────────────────────
    3632:  {"name": "distccd",       "reason": "Distributed compiler daemon. Allows arbitrary code execution (CVE-2004-2687)."},
}

INTERESTING_PORTS = {
    # ── Remote access ─────────────────────────────────────────────────────
    22:    {"name": "SSH",           "reason": "Remote access. Check version, authentication methods, and root login."},
    3389:  {"name": "RDP",           "reason": "Windows remote desktop. Frequent brute force and BlueKeep target."},
    5900:  {"name": "VNC",           "reason": "Remote desktop. Frequently weak or no authentication."},
    5901:  {"name": "VNC-1",         "reason": "VNC alternate port. Same risks : verify authentication."},
    # ── Windows services ─────────────────────────────────────────────────
    445:   {"name": "SMB",           "reason": "Windows file sharing. Critical CVE history (EternalBlue, PrintNightmare)."},
    139:   {"name": "NetBIOS-SSN",   "reason": "SMB over NetBIOS. Same risk as port 445."},
    5985:  {"name": "WinRM HTTP",    "reason": "Windows remote management. May allow command execution."},
    5986:  {"name": "WinRM HTTPS",   "reason": "Windows remote management (encrypted). Verify auth."},
    # ── Directory services / auth ─────────────────────────────────────────
    389:   {"name": "LDAP",          "reason": "Active directory. User and object enumeration possible."},
    636:   {"name": "LDAPS",         "reason": "LDAP over SSL. Verify anonymous query permissions."},
    88:    {"name": "Kerberos",      "reason": "Kerberos auth. AS-REP Roasting, Kerberoasting possible."},
    3268:  {"name": "GlobalCatalog", "reason": "AD Global Catalog LDAP. Full domain enumeration possible."},
    3269:  {"name": "GlobalCatalog SSL","reason": "AD Global Catalog LDAPS. Verify query permissions."},
    # ── Databases ────────────────────────────────────────────────────────
    1433:  {"name": "MSSQL",         "reason": "SQL Server. Should not be externally exposed."},
    3306:  {"name": "MySQL",         "reason": "MySQL database. Should not be externally exposed."},
    5432:  {"name": "PostgreSQL",    "reason": "PostgreSQL database. Should not be externally exposed."},
    1521:  {"name": "Oracle DB",     "reason": "Oracle database. Should not be externally exposed."},
    1527:  {"name": "Oracle DB alt", "reason": "Oracle alternate port. Should not be externally exposed."},
    6379:  {"name": "Redis",         "reason": "Redis has no auth by default. Possible RCE via config commands."},
    27017: {"name": "MongoDB",       "reason": "MongoDB no auth by default in old versions."},
    27018: {"name": "MongoDB (shard)","reason": "MongoDB shard port. Verify authentication."},
    9200:  {"name": "Elasticsearch", "reason": "No auth by default. Massive data exposure risk."},
    9300:  {"name": "Elasticsearch Cluster","reason": "ES cluster comms. Internal traffic that should not be exposed."},
    11211: {"name": "Memcached",     "reason": "No authentication. DDoS amplification attack vector."},
    5984:  {"name": "CouchDB",       "reason": "CouchDB. Admin party mode (no auth) in old versions."},
    6432:  {"name": "PgBouncer",     "reason": "PostgreSQL connection pooler. Verify auth configuration."},
    # ── Web ───────────────────────────────────────────────────────────────
    80:    {"name": "HTTP",          "reason": "Unencrypted web traffic. Verify content and redirect to HTTPS."},
    443:   {"name": "HTTPS",         "reason": "Encrypted web. Verify certificate and application security."},
    8080:  {"name": "HTTP-Alt",      "reason": "Alternative web. Common for admin panels and dev servers."},
    8443:  {"name": "HTTPS-Alt",     "reason": "Alternative HTTPS. Verify certificate and auth."},
    8000:  {"name": "HTTP Dev",      "reason": "Common dev server port. Verify it is not exposed to production."},
    8888:  {"name": "HTTP Dev",      "reason": "Common for Jupyter Notebooks (no auth by default)."},
    # ── Mail ─────────────────────────────────────────────────────────────
    25:    {"name": "SMTP",          "reason": "Mail. Check open relay and user enumeration (VRFY/EXPN)."},
    465:   {"name": "SMTPS",         "reason": "SMTP over SSL. Verify relay configuration."},
    587:   {"name": "SMTP Submission","reason": "Mail submission. Check for open relay."},
    110:   {"name": "POP3",          "reason": "Unencrypted mail retrieval. Credentials exposed."},
    995:   {"name": "POP3S",         "reason": "POP3 over SSL. Verify certificate."},
    143:   {"name": "IMAP",          "reason": "Unencrypted mail access. Credentials exposed."},
    993:   {"name": "IMAPS",         "reason": "IMAP over SSL. Verify certificate."},
    # ── DNS ───────────────────────────────────────────────────────────────
    53:    {"name": "DNS",           "reason": "DNS server. Check zone transfers (AXFR) and recursion."},
    # ── Container / Cloud ─────────────────────────────────────────────────
    2375:  {"name": "Docker API",    "reason": "Docker API without TLS. Full host control if exposed."},
    2376:  {"name": "Docker TLS",    "reason": "Docker API with TLS. Verify certificate configuration."},
    6443:  {"name": "Kubernetes API","reason": "K8s API. Cluster control if misconfigured."},
    2379:  {"name": "etcd",          "reason": "Kubernetes etcd. Full cluster state and secrets exposed."},
    2380:  {"name": "etcd cluster",  "reason": "etcd cluster peer port. Should be internal only."},
    10250: {"name": "Kubelet API",   "reason": "Kubelet. Can allow exec into pods and node-level actions."},
    # ── Monitoring / observability ────────────────────────────────────────
    9090:  {"name": "Prometheus",    "reason": "Exposed system metrics. Sensitive infrastructure info."},
    3000:  {"name": "Grafana/Dev",   "reason": "Common for Grafana. Default admin:admin credentials."},
    9100:  {"name": "Node Exporter", "reason": "Prometheus Node Exporter. Full system metrics exposed."},
    # ── Application servers ───────────────────────────────────────────────
    7001:  {"name": "WebLogic",      "reason": "Oracle WebLogic. Multiple critical RCE CVEs."},
    7002:  {"name": "WebLogic SSL",  "reason": "WebLogic HTTPS. Same risks : verify patching."},
    8161:  {"name": "ActiveMQ",      "reason": "ActiveMQ Admin. Documented RCE vulnerabilities."},
    61616: {"name": "ActiveMQ Broker","reason": "ActiveMQ message broker. Deserialization RCE possible."},
    4848:  {"name": "GlassFish",     "reason": "GlassFish Admin. History of critical CVEs."},
    9000:  {"name": "SonarQube/PHP-FPM","reason": "Common for SonarQube or PHP-FPM. Verify access."},
    8009:  {"name": "AJP",           "reason": "Apache JServ Protocol. Ghostcat vulnerability (CVE-2020-1938)."},
    # ── VPN / tunneling ───────────────────────────────────────────────────
    500:   {"name": "IKE (IPSec)",   "reason": "IPSec key exchange. Verify IKEv2 and strong ciphers."},
    1194:  {"name": "OpenVPN",       "reason": "OpenVPN server. Verify certificate and auth configuration."},
    1723:  {"name": "PPTP VPN",      "reason": "PPTP VPN. Weak protocol : MS-CHAPv2 is crackable."},
    4500:  {"name": "IPSec NAT-T",   "reason": "IPSec NAT traversal. Verify configuration."},
    # ── Misc high-risk ────────────────────────────────────────────────────
    5000:  {"name": "UPnP/Flask Dev","reason": "Common for UPnP or Flask dev server. Verify what is running."},
    8500:  {"name": "Consul",        "reason": "HashiCorp Consul. Service mesh : full cluster info if exposed."},
    8200:  {"name": "Vault",         "reason": "HashiCorp Vault. Secret management : should never be externally exposed."},
    5601:  {"name": "Kibana",        "reason": "Kibana dashboard. Full Elasticsearch data access."},
    15672: {"name": "RabbitMQ Mgmt","reason": "RabbitMQ management UI. Default guest:guest credentials."},
    5672:  {"name": "AMQP",          "reason": "RabbitMQ/AMQP broker. Message interception possible."},
}

# ─────────────────────────────────────────────────────────────────────────────
# FOLLOW-UP RECOMMENDATIONS (when port is open but no service detection)
# ─────────────────────────────────────────────────────────────────────────────

FOLLOWUP_COMMANDS = {
    445:   "nmap -sV -p 445 --script smb-security-mode,smb2-security-mode,smb-vuln-ms17-010,smb-vuln-ms08-067 {ip}",
    139:   "nmap -sV -p 139,445 --script smb-security-mode,smb-enum-shares {ip}",
    22:    "nmap -sV -p 22 --script ssh-auth-methods,ssh-hostkey,ssh2-enum-algos {ip}",
    21:    "nmap -sV -p 21 --script ftp-anon,ftp-bounce,ftp-syst {ip}",
    23:    "nmap -sV -p 23 --script telnet-encryption,telnet-ntlm-info {ip}",
    25:    "nmap -sV -p 25 --script smtp-open-relay,smtp-enum-users,smtp-commands {ip}",
    53:    "nmap -sV -p 53 --script dns-zone-transfer,dns-recursion {ip}",
    80:    "nmap -sV -p 80 --script http-title,http-headers,http-methods,http-auth-finder {ip}",
    443:   "nmap -sV -p 443 --script ssl-cert,ssl-enum-ciphers,ssl-heartbleed,http-title {ip}",
    161:   "nmap -sU -p 161 --script snmp-brute,snmp-info,snmp-sysdescr {ip}",
    3306:  "nmap -sV -p 3306 --script mysql-empty-password,mysql-info {ip}",
    1433:  "nmap -sV -p 1433 --script ms-sql-info,ms-sql-empty-password {ip}",
    3389:  "nmap -sV -p 3389 --script rdp-enum-encryption,rdp-vuln-ms12-020 {ip}",
    5900:  "nmap -sV -p 5900 --script vnc-info,vnc-brute {ip}",
    6379:  "nmap -sV -p 6379 --script redis-info {ip}",
    27017: "nmap -sV -p 27017 --script mongodb-info {ip}",
    5432:  "nmap -sV -p 5432 --script pgsql-brute {ip}",
    9200:  "nmap -sV -p 9200 --script http-title {ip}",
    2375:  "nmap -sV -p 2375 --script http-title {ip}",
    623:   "nmap -sU -p 623 --script ipmi-version,ipmi-brute {ip}",
    8080:  "nmap -sV -p 8080 --script http-title,http-headers,http-auth-finder {ip}",
    8443:  "nmap -sV -p 8443 --script ssl-cert,ssl-enum-ciphers,http-title {ip}",
    5985:  "nmap -sV -p 5985 --script http-auth-finder {ip}",
}

GENERIC_FOLLOWUP = "nmap -sV -p {port} --script default {ip}"

# ─────────────────────────────────────────────────────────────────────────────
# VULNERABILITY DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def detect_vulnerabilities(host_data):
    vulns = []
    for port in host_data.get('ports', []):
        port_num = port.get('portid', 0)
        service  = port.get('service', {})
        svc_name = service.get('name', '').lower()
        product  = service.get('product', '').lower()
        version  = service.get('version', '').lower()
        scripts  = port.get('scripts', [])
        full_ver = f"{product} {version}".strip()
        method   = service.get('method', 'table')

        is_critical    = port_num in CRITICAL_PORTS
        is_interesting = port_num in INTERESTING_PORTS
        port_info      = CRITICAL_PORTS.get(port_num) or INTERESTING_PORTS.get(port_num)

        # ── No service info at all: generate a TODO, not a vulnerability ───
        # These are actionable investigation tasks, not confirmed findings.
        if not service.get('name') and not service.get('product'):
            if is_critical or is_interesting:
                cmd = FOLLOWUP_COMMANDS.get(port_num, GENERIC_FOLLOWUP).format(
                    ip=host_data.get('ip', '[TARGET]'), port=port_num)
                pname = port_info['name']
                host_data.setdefault('todos', []).append({
                    "port": port_num,
                    "priority": "high" if is_critical else "medium",
                    "title": f"Investigate {pname} on port {port_num}",
                    "reason": port_info['reason'],
                    "description": (
                        f"Port {port_num} is open. It is typically used by {pname} "
                        f"but no service banner was captured (scan ran without -sV). "
                        f"Confirm what is running and assess for vulnerabilities."
                    ),
                    "followup": cmd,
                })
            # Any other port with no service info: skip silently
            continue

        # ── Service identified by port number only (table match, no -sV) ──
        # Generate a TODO for critical/interesting ports, not a vulnerability.
        if method == 'table' and (is_critical or is_interesting):
            cmd = FOLLOWUP_COMMANDS.get(port_num, GENERIC_FOLLOWUP).format(
                ip=host_data.get('ip', '[TARGET]'), port=port_num)
            pname = port_info['name']
            host_data.setdefault('todos', []).append({
                "port": port_num,
                "priority": "high" if is_critical else "medium",
                "title": f"Run version scan on {pname} (port {port_num})",
                "reason": port_info['reason'],
                "description": (
                    f"{pname} identified on port {port_num} by port number only. "
                    f"No service banner was read (scan ran without -sV). "
                    f"Version info is required to detect known CVEs and misconfigurations."
                ),
                "followup": cmd,
            })

        # ── SEVERITY FRAMEWORK ────────────────────────────────────────────
        # CRITICAL : Confirmed CVEs with RCE/authentication bypass, or cleartext
        #            legacy protocols that must never be exposed under any circumstance.
        # HIGH     : Services/configs that are dangerous when internet-exposed or
        #            reachable by untrusted users; may enable lateral movement.
        # MEDIUM   : Misconfigurations or weak settings that require a deliberate
        #            attack or insider position to exploit.
        # LOW      : Best-practice deviations with limited direct impact.
        # INFO     : Informational observations with no direct exploitability.

        # ── SMB ───────────────────────────────────────────────────────────
        if port_num in (445, 139) or 'smb' in svc_name:
            for script in scripts:
                output = script.get('output', '').lower()
                sid    = script.get('id', '').lower()
                if 'smb-security-mode' in sid:
                    if 'message_signing: disabled' in output or 'signing: disabled' in output:
                        # HIGH: Enables relay attacks but requires network position
                        vulns.append({"severity": "HIGH", "port": port_num,
                            "title": "SMB Signing Disabled",
                            "description": "SMB without digital signing. Enables NTLM Relay / Pass-the-Hash attacks from the same network segment.",
                            "cve": "N/A", "script": sid})
                    if 'account_used: guest' in output or 'anonymous' in output:
                        # CRITICAL: Unauthenticated access to file shares
                        vulns.append({"severity": "CRITICAL", "port": port_num,
                            "title": "SMB Anonymous / Guest Access Enabled",
                            "description": "SMB allows unauthenticated access. Attackers can enumerate shares, read files and potentially write without credentials.",
                            "cve": "N/A", "script": sid})
                if 'smb-vuln-ms17-010' in sid or 'eternalblue' in output:
                    # CRITICAL: confirmed wormable RCE CVE
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": "EternalBlue (MS17-010) - Wormable RCE",
                        "description": "Host is vulnerable to EternalBlue. Unauthenticated RCE as SYSTEM. Exploited by WannaCry and NotPetya. Patch immediately.",
                        "cve": "CVE-2017-0144", "script": sid})
                if 'smb-vuln-ms08-067' in sid:
                    # CRITICAL: confirmed unauthenticated RCE CVE
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": "MS08-067 NetAPI Unauthenticated RCE",
                        "description": "Unauthenticated remote code execution on Windows XP/2003. Highly reliable exploit available.",
                        "cve": "CVE-2008-4250", "script": sid})
                if 'smb2-security-mode' in sid and 'signing enabled and required' not in output:
                    # MEDIUM: requires network position, not directly exploitable alone
                    vulns.append({"severity": "MEDIUM", "port": port_num,
                        "title": "SMBv2 Signing Not Required",
                        "description": "SMBv2 signing is not enforced. In combination with credential capture (Responder, etc.), relay attacks are possible.",
                        "cve": "N/A", "script": sid})
            if 'smbv1' in full_ver:
                # CRITICAL: obsolete protocol, direct CVE vector
                vulns.append({"severity": "CRITICAL", "port": port_num,
                    "title": "SMBv1 Protocol Enabled",
                    "description": "SMBv1 is active. This deprecated protocol is the attack surface for EternalBlue/WannaCry. Disable it in Windows features.",
                    "cve": "CVE-2017-0144", "script": "version-detection"})

        # ── SSH ───────────────────────────────────────────────────────────
        if 'ssh' in svc_name or port_num == 22:
            for script in scripts:
                sid    = script.get('id', '').lower()
                output = script.get('output', '').lower()
                if 'ssh-auth-methods' in sid and 'password' in output:
                    # LOW: password auth is common, exploitable only with valid credentials
                    vulns.append({"severity": "LOW", "port": port_num,
                        "title": "SSH Password Authentication Enabled",
                        "description": "SSH accepts password-based login. Consider enforcing key-only authentication to reduce brute-force exposure.",
                        "cve": "N/A", "script": sid})
                if 'ssh-hostkey' in sid and 'rsa 1024' in output:
                    # MEDIUM: weak key, practical attack requires significant resources
                    vulns.append({"severity": "MEDIUM", "port": port_num,
                        "title": "SSH Host Key Too Short (RSA 1024-bit)",
                        "description": "1024-bit RSA host keys no longer meet modern security standards. Regenerate with a minimum of 2048 bits (3072 recommended).",
                        "cve": "N/A", "script": sid})
            if version and any(v in version for v in ['openssh 4.', 'openssh 5.', 'openssh 6.']):
                # HIGH: many public CVEs in legacy OpenSSH, some pre-auth
                vulns.append({"severity": "HIGH", "port": port_num,
                    "title": f"Outdated OpenSSH Version ({version})",
                    "description": "This OpenSSH release has numerous public CVEs including pre-authentication vulnerabilities. Update to the latest stable release.",
                    "cve": "Multiple CVEs", "script": "version-detection"})

        # ── SSL/TLS ───────────────────────────────────────────────────────
        for script in scripts:
            sid    = script.get('id', '').lower()
            output = script.get('output', '').lower()
            if 'ssl-heartbleed' in sid and 'vulnerable' in output:
                # CRITICAL: confirmed CVE, private key disclosure
                vulns.append({"severity": "CRITICAL", "port": port_num,
                    "title": "Heartbleed (OpenSSL Memory Disclosure)",
                    "description": "Confirmed vulnerable to Heartbleed. An unauthenticated attacker can read server memory, potentially recovering private keys, session tokens and passwords.",
                    "cve": "CVE-2014-0160", "script": sid})
            if 'ssl-poodle' in sid and 'vulnerable' in output:
                # HIGH: requires MitM position, but SSLv3 should never be enabled
                vulns.append({"severity": "HIGH", "port": port_num,
                    "title": "POODLE - SSLv3 Accepted",
                    "description": "Server accepts SSLv3, which is fundamentally broken. A network attacker can decrypt HTTPS sessions via the POODLE downgrade attack. Disable SSLv3.",
                    "cve": "CVE-2014-3566", "script": sid})
            if 'ssl-drown' in sid and 'vulnerable' in output:
                # CRITICAL: cross-protocol decryption, confirmed CVE
                vulns.append({"severity": "CRITICAL", "port": port_num,
                    "title": "DROWN Attack - SSLv2 Accepted",
                    "description": "Server supports SSLv2, enabling the DROWN attack which can decrypt TLS sessions. Any server sharing the same private key is at risk.",
                    "cve": "CVE-2016-0800", "script": sid})
            if 'ssl-cert' in sid and 'expired' in output:
                # LOW: no direct exploitability, but breaks trust chain
                vulns.append({"severity": "LOW", "port": port_num,
                    "title": "Expired TLS Certificate",
                    "description": "The server's TLS certificate is expired. Clients receive browser warnings and encrypted channels may not be trusted. Renew the certificate.",
                    "cve": "N/A", "script": sid})

        # ── FTP ───────────────────────────────────────────────────────────
        if 'ftp' in svc_name or port_num == 21:
            for script in scripts:
                sid    = script.get('id', '').lower()
                output = script.get('output', '').lower()
                if 'ftp-anon' in sid and ('allowed' in output or 'anonymous' in output):
                    # HIGH: unauthenticated file system access
                    vulns.append({"severity": "HIGH", "port": port_num,
                        "title": "FTP Anonymous Access Allowed",
                        "description": "FTP server permits anonymous login. Depending on write permissions, attackers can exfiltrate files or upload malicious content without credentials.",
                        "cve": "N/A", "script": sid})
            # FTP itself is cleartext — only flag if confirmed running (probed method)
            if method == 'probed' and 'ftp' in svc_name:
                vulns.append({"severity": "HIGH", "port": port_num,
                    "title": "FTP Service Exposed (Cleartext Protocol)",
                    "description": "FTP transmits credentials and data in cleartext. All traffic is interceptable. Replace with SFTP or FTPS if remote file transfer is required.",
                    "cve": "N/A", "script": "version-detection"})

        # ── Redis ─────────────────────────────────────────────────────────
        if 'redis' in svc_name or port_num == 6379:
            for script in scripts:
                if 'redis-info' in script.get('id', '').lower():
                    # CRITICAL: unauthenticated access, confirmed reachable
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": "Redis Accessible Without Authentication",
                        "description": "Redis responded without credentials. Attackers can read/write all data and in many configurations achieve RCE (via SLAVEOF or config rewrite).",
                        "cve": "N/A", "script": script['id']})

        # ── Databases exposed without authentication evidence ──────────────
        DB_PORTS = {
            3306: 'MySQL', 5432: 'PostgreSQL', 1433: 'MSSQL', 1521: 'Oracle',
            27017: 'MongoDB', 5984: 'CouchDB', 9200: 'Elasticsearch',
        }
        if port_num in DB_PORTS and method == 'probed':
            db_name = DB_PORTS[port_num]
            for script in scripts:
                sid = script.get('id','').lower()
                out = script.get('output','').lower()
                if 'empty-password' in sid and ('ok' in out or 'success' in out):
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": f"{db_name}: Empty/No Password",
                        "description": f"{db_name} accepts connections with an empty or default password. Immediate credential remediation required.",
                        "cve": "N/A", "script": sid})
                elif 'info' in sid or 'brute' in sid:
                    vulns.append({"severity": "HIGH", "port": port_num,
                        "title": f"{db_name} Accessible from Network",
                        "description": f"{db_name} (port {port_num}) responds to external connections. Databases should not be internet-reachable. Restrict access to application servers only.",
                        "cve": "N/A", "script": sid})

        # ── HTTP/Web ──────────────────────────────────────────────────────
        if svc_name in ('http','https','http-alt','https-alt') or port_num in (80,443,8080,8443):
            for script in scripts:
                sid    = script.get('id', '').lower()
                output = script.get('output', '').lower()
                if 'http-shellshock' in sid and 'vulnerable' in output:
                    # CRITICAL: confirmed unauthenticated RCE via CGI
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": "Shellshock - Remote Code Execution via CGI",
                        "description": "Web server confirmed vulnerable to Shellshock. Unauthenticated RCE is possible via specially crafted HTTP headers against any CGI script.",
                        "cve": "CVE-2014-6271", "script": sid})
                if 'http-default-accounts' in sid and 'valid' in output:
                    # CRITICAL: confirmed working default credentials
                    vulns.append({"severity": "CRITICAL", "port": port_num,
                        "title": "Default Credentials Accepted",
                        "description": "The web application accepts known default credentials. This gives an attacker authenticated access. Change all default passwords immediately.",
                        "cve": "N/A", "script": sid})
                if 'http-open-redirect' in sid and 'found' in output:
                    vulns.append({"severity": "LOW", "port": port_num,
                        "title": "Open Redirect Detected",
                        "description": "The web application redirects to external URLs without validation, enabling phishing attacks.",
                        "cve": "N/A", "script": sid})

        # ── RDP ───────────────────────────────────────────────────────────
        if port_num == 3389 or 'rdp' in svc_name or 'ms-wbt' in svc_name:
            for script in scripts:
                sid    = script.get('id', '').lower()
                output = script.get('output', '').lower()
                if 'rdp-vuln-ms12-020' in sid and 'vulnerable' in output:
                    # HIGH: confirmed CVE, DoS / potential memory corruption
                    vulns.append({"severity": "HIGH", "port": port_num,
                        "title": "MS12-020 RDP Denial-of-Service",
                        "description": "RDP is vulnerable to MS12-020, allowing an unauthenticated attacker to crash the system via malformed packets.",
                        "cve": "CVE-2012-0152", "script": sid})
                if 'rdp-enum-encryption' in sid and 'rdp security layer' in output:
                    # MEDIUM: weaker auth, not directly exploitable but increases risk
                    vulns.append({"severity": "MEDIUM", "port": port_num,
                        "title": "RDP Without Network Level Authentication (NLA)",
                        "description": "RDP uses legacy encryption instead of NLA. Authentication occurs after a full connection, enabling credential brute-force without prior authentication.",
                        "cve": "N/A", "script": sid})
            # Exposed RDP is inherently high-risk on internet-facing hosts
            if method == 'probed':
                vulns.append({"severity": "HIGH", "port": port_num,
                    "title": "RDP Exposed (Remote Desktop Protocol)",
                    "description": "RDP (port 3389) is reachable. Internet-facing RDP is a top ransomware entry point. Restrict to VPN or trusted IPs only.",
                    "cve": "N/A", "script": "port-classification"})

        # ── Telnet ────────────────────────────────────────────────────────
        if 'telnet' in svc_name or port_num == 23:
            # CRITICAL: cleartext credentials, universally deprecated
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "Telnet Service Active (Cleartext Remote Access)",
                "description": "Telnet transmits all data including credentials in cleartext. Any network observer can capture login sessions. Replace with SSH immediately.",
                "cve": "N/A", "script": "port-classification"})

        # ── IPMI / BMC ────────────────────────────────────────────────────
        if port_num == 623:
            # CRITICAL: confirmed CVE with authentication bypass
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "IPMI / BMC Exposed (Auth Bypass CVE)",
                "description": "IPMI is reachable. Cipher 0 allows authentication bypass (CVE-2013-4786), enabling attackers to retrieve password hashes and gain full out-of-band server control.",
                "cve": "CVE-2013-4786", "script": "port-classification"})

        # ── Cisco Smart Install ────────────────────────────────────────────
        if port_num == 4786:
            # CRITICAL: confirmed unauthenticated RCE on network devices
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "Cisco Smart Install - Unauthenticated RCE",
                "description": "Smart Install protocol is exposed. No authentication is required. Attackers can overwrite device configurations, change IOS images, and execute arbitrary commands.",
                "cve": "CVE-2018-0171", "script": "port-classification"})

        # ── AJP / Ghostcat ────────────────────────────────────────────────
        if port_num == 8009 or 'ajp' in svc_name:
            # CRITICAL: confirmed CVE, file read / potential RCE
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "AJP Connector Exposed (Ghostcat - CVE-2020-1938)",
                "description": "The Apache JServ Protocol port is reachable. Ghostcat allows unauthenticated reading of any file within the web application and code execution if file upload is available.",
                "cve": "CVE-2020-1938", "script": "port-classification"})

        # ── Docker / Kubernetes APIs ───────────────────────────────────────
        if port_num == 2375:
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "Docker API Exposed Without TLS",
                "description": "The Docker daemon API is reachable without TLS. An attacker gains full container orchestration control and can escape to the host OS trivially.",
                "cve": "N/A", "script": "port-classification"})
        if port_num == 2379:
            vulns.append({"severity": "CRITICAL", "port": port_num,
                "title": "etcd Exposed (Kubernetes Secrets Store)",
                "description": "etcd is reachable. It stores all Kubernetes secrets, tokens and configuration in plaintext. An unauthenticated read gives full cluster credential access.",
                "cve": "N/A", "script": "port-classification"})

        # ── Outdated software versions ─────────────────────────────────────
        ver_patterns = [
            # (regex, friendly_name, severity, reason)
            (r'apache[/ ](1\.\d+)', 'Apache 1.x', 'CRITICAL',
             'Apache 1.x is end-of-life with multiple unpatched RCE and DoS CVEs.'),
            (r'apache[/ ](2\.[0-3])', 'Apache 2.x (old)', 'HIGH',
             'Old Apache 2.x branch with multiple publicly known CVEs. Update to 2.4.x latest.'),
            (r'nginx[/ ](0\.\d+|1\.[0-9]\.)', 'Nginx (legacy)', 'HIGH',
             'Legacy Nginx version with known vulnerabilities. Update to current stable.'),
            (r'php[/ ](4\.\d+|5\.\d+)', 'PHP 4/5', 'CRITICAL',
             'PHP 4 and 5 are end-of-life with hundreds of unpatched CVEs including RCE.'),
            (r'php[/ ](7\.[01])', 'PHP 7.0/7.1', 'HIGH',
             'PHP 7.0 and 7.1 are end-of-life. Multiple CVEs including type confusion bugs.'),
            (r'iis[/ ]([456]\.\d+)', 'IIS legacy', 'CRITICAL',
             'IIS 4-6 are end-of-life with critical publicly exploited CVEs.'),
            (r'iis[/ ](7\.\d+)', 'IIS 7.x', 'HIGH',
             'IIS 7.x receives no security updates. Migrate to IIS 10.'),
            (r'openssh[/ ](4\.|5\.|6\.)', 'OpenSSH legacy', 'HIGH',
             'Very old OpenSSH with multiple pre-auth vulnerabilities. Update immediately.'),
            (r'openssl[/ ](0\.|1\.0\.[01])', 'OpenSSL legacy', 'CRITICAL',
             'End-of-life OpenSSL with critical CVEs (Heartbleed range). Update to 3.x.'),
        ]
        for pattern, svc_friendly, sev, reason in ver_patterns:
            m = re.search(pattern, full_ver, re.IGNORECASE)
            if m:
                vulns.append({"severity": sev, "port": port_num,
                    "title": f"End-of-Life / Vulnerable Version: {svc_friendly}",
                    "description": reason,
                    "cve": "Multiple CVEs", "script": "version-detection"})
                break  # only one version finding per port

    return vulns


# ─────────────────────────────────────────────────────────────────────────────
# NMAP XML PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_nmap_xml(xml_content):
    root = ET.fromstring(xml_content)
    scan_info = {
        "scanner":     root.get('scanner', 'nmap'),
        "args":        root.get('args', ''),
        "start":       root.get('startstr', ''),
        "version":     root.get('version', ''),
        "has_version": False,
        "has_scripts": False,
        "hosts":       []
    }
    args = scan_info['args'].lower()
    if any(f in args for f in ['-sv', '-a ', '-sc', '--script']):
        scan_info['has_version'] = True
    if any(f in args for f in ['--script', '-sc', '-a ']):
        scan_info['has_scripts'] = True

    for host in root.findall('host'):
        status = host.find('status')
        if status is None or status.get('state') != 'up':
            continue
        host_data = {"ip": "", "hostname": "", "os": "", "os_accuracy": 0,
                     "mac": "", "vendor": "", "state": "up", "ports": [], "vulns": [], "todos": []}

        for addr in host.findall('address'):
            t = addr.get('addrtype')
            if t == 'ipv4':   host_data['ip'] = addr.get('addr', '')
            elif t == 'ipv6' and not host_data['ip']: host_data['ip'] = addr.get('addr', '')
            elif t == 'mac':
                host_data['mac']    = addr.get('addr', '')
                host_data['vendor'] = addr.get('vendor', '')

        hostnames_el = host.find('hostnames')
        if hostnames_el is not None:
            hn = hostnames_el.find('hostname')
            if hn is not None:
                host_data['hostname'] = hn.get('name', '')

        os_el = host.find('os')
        if os_el is not None:
            for match in os_el.findall('osmatch'):
                acc = int(match.get('accuracy', 0))
                if acc > host_data['os_accuracy']:
                    host_data['os']          = match.get('name', '')
                    host_data['os_accuracy'] = acc

        ports_el = host.find('ports')
        if ports_el is not None:
            for port_el in ports_el.findall('port'):
                state_el = port_el.find('state')
                if state_el is None or state_el.get('state') != 'open':
                    continue
                port_num   = int(port_el.get('portid', 0))
                protocol   = port_el.get('protocol', 'tcp')
                service_el = port_el.find('service')
                service    = {}
                if service_el is not None:
                    service = {
                        "name":      service_el.get('name', ''),
                        "product":   service_el.get('product', ''),
                        "version":   service_el.get('version', ''),
                        "extrainfo": service_el.get('extrainfo', ''),
                        "method":    service_el.get('method', 'table'),
                    }
                    if service.get('method') == 'probed':
                        scan_info['has_version'] = True
                scripts = []
                for s in port_el.findall('script'):
                    scripts.append({"id": s.get('id',''), "output": s.get('output','')})
                if scripts:
                    scan_info['has_scripts'] = True

                classification, class_reason = "normal", ""
                if port_num in CRITICAL_PORTS:
                    classification = "critical"
                    class_reason   = CRITICAL_PORTS[port_num]['reason']
                elif port_num in INTERESTING_PORTS:
                    classification = "interesting"
                    class_reason   = INTERESTING_PORTS[port_num]['reason']

                host_data['ports'].append({
                    "portid": port_num, "protocol": protocol,
                    "service": service, "scripts": scripts,
                    "classification": classification, "class_reason": class_reason,
                })
        host_data['vulns'] = detect_vulnerabilities(host_data)
        # Skip hosts with no open ports — they clutter the report without adding value
        if host_data['ports']:
            scan_info['hosts'].append(host_data)
    return scan_info


def merge_scans(scan_list):
    merged = {
        "scanner":     "nmap",
        "args":        " | ".join(s['args'] for s in scan_list if s.get('args')),
        "start":       scan_list[0].get('start', '') if scan_list else '',
        "version":     scan_list[0].get('version', '') if scan_list else '',
        "has_version": any(s.get('has_version') for s in scan_list),
        "has_scripts": any(s.get('has_scripts') for s in scan_list),
        "hosts":       []
    }
    hosts_by_ip = {}
    for scan in scan_list:
        for host in scan.get('hosts', []):
            ip = host['ip']
            if ip not in hosts_by_ip:
                hosts_by_ip[ip] = host
            else:
                existing = hosts_by_ip[ip]
                existing_ports = {p['portid']: p for p in existing['ports']}
                for port in host['ports']:
                    pid = port['portid']
                    if pid not in existing_ports:
                        existing_ports[pid] = port
                    else:
                        ep = existing_ports[pid]
                        if not ep['service'].get('product') and port['service'].get('product'):
                            existing_ports[pid] = port
                        existing_script_ids = {s['id'] for s in ep.get('scripts', [])}
                        for script in port.get('scripts', []):
                            if script['id'] not in existing_script_ids:
                                ep.setdefault('scripts', []).append(script)
                existing['ports'] = list(existing_ports.values())
                if host.get('os_accuracy', 0) > existing.get('os_accuracy', 0):
                    existing['os']          = host['os']
                    existing['os_accuracy'] = host['os_accuracy']
                existing['vulns'] = detect_vulnerabilities(existing)
    merged['hosts'] = list(hosts_by_ip.values())
    return merged


# ─────────────────────────────────────────────────────────────────────────────
# HISTORY
# ─────────────────────────────────────────────────────────────────────────────

def save_to_history(scan_data, filenames):
    ts      = datetime.now()
    scan_id = ts.strftime('%Y%m%d_%H%M%S')
    total_vulns = sum(len(h.get('vulns', [])) for h in scan_data.get('hosts', []))
    crit_ports  = sum(
        len([p for p in h.get('ports', []) if p.get('classification') == 'critical'])
        for h in scan_data.get('hosts', []))
    meta = {
        "id":                scan_id,
        "timestamp":         ts.isoformat(),
        "timestamp_display": ts.strftime('%Y-%m-%d %H:%M:%S'),
        "filenames":         filenames,
        "host_count":        len(scan_data.get('hosts', [])),
        "total_ports":       sum(len(h.get('ports', [])) for h in scan_data.get('hosts', [])),
        "vuln_count":        total_vulns,
        "critical_ports":    crit_ports,
        "args":              scan_data.get('args', ''),
    }
    with open(HISTORY_DIR / f"{scan_id}.json", 'w') as f:
        json.dump({**meta, "scan": scan_data}, f, indent=2)
    return meta


# ─────────────────────────────────────────────────────────────────────────────
# REPORT: HTML
# ─────────────────────────────────────────────────────────────────────────────

def generate_html_report(scan, meta):
    ts    = meta.get('timestamp_display', '')
    files = ', '.join(meta.get('filenames', []))
    sev_color = {'CRITICAL':'#ff3333','HIGH':'#ff8c00','MEDIUM':'#ffd700','LOW':'#00bfff','INFO':'#888'}
    hosts_html = ""
    for host in scan.get('hosts', []):
        sorted_ports = sorted(host.get('ports',[]),
                              key=lambda p:{'critical':0,'interesting':1,'normal':2}.get(p['classification'],2))
        pts_parts = []
        for p in sorted_ports:
            svc = p.get('service', {})
            icon = '🔴' if p['classification']=='critical' else '🟠' if p['classification']=='interesting' else '🟢'
            ver = ' '.join(filter(None, [svc.get('product',''), svc.get('version','')]))
            pts_parts.append(
                "<tr><td>" + icon + " " + str(p['portid']) + "/" + p['protocol'] + "</td>"
                + "<td>" + svc.get('name','') + "</td>"
                + "<td>" + ver + "</td>"
                + "<td style='color:#888;font-size:11px'>" + p.get('class_reason','') + "</td></tr>"
            )
        pts = "".join(pts_parts)
        sev_order = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFO':4}
        vls_parts = []
        for v in sorted(host.get('vulns',[]), key=lambda x: sev_order.get(x.get('severity','INFO'), 9)):
            col = sev_color.get(v['severity'], '#888')
            vls_parts.append(
                "<tr><td style='color:" + col + ";font-weight:700'>" + v['severity'] + "</td>"
                + "<td>" + v['title'] + "</td>"
                + "<td>" + v['description'] + "</td>"
                + "<td style='color:#bc8cff'>" + v.get('cve','N/A') + "</td>"
                + "<td style='color:#58a6ff'>" + str(v['port']) + "</td></tr>"
            )
        vls = "".join(vls_parts)
        todos_html = ""
        if host.get('todos'):
            todo_rows = "".join(
                "<tr><td style='color:" + ('#bc8cff' if t.get('priority')=='high' else '#e07a30') + ";font-weight:700'>"
                + t.get('priority','').upper() + "</td>"
                + "<td>" + str(t.get('port','')) + "</td>"
                + "<td>" + t.get('title','') + "</td>"
                + "<td style='color:#888;font-size:11px'>" + t.get('description','') + "</td>"
                + "<td style='color:#00bfff;font-size:10px;font-family:monospace'>" + (t.get('followup','') or '') + "</td></tr>"
                for t in host.get('todos',[])
            )
            todos_html = f"<h4>INVESTIGATION ACTIONS ({len(host.get('todos',[]))})</h4><table><tr><th>Priority</th><th>Port</th><th>Title</th><th>Description</th><th>Follow-up Scan</th></tr>{todo_rows}</table>"
        hosts_html += f"""<div class='hb'>
          <h3>// {host.get('ip','')} {('<span style="color:#888">'+host['hostname']+'</span>') if host.get('hostname') else ''}</h3>
          <p class='meta'>OS: {host.get('os','Unknown')} | MAC: {host.get('mac','N/A')} {host.get('vendor','')}</p>
          <h4>OPEN PORTS ({len(host.get('ports',[]))})</h4>
          <table><tr><th>Port</th><th>Service</th><th>Version</th><th>Note</th></tr>{pts}</table>
          {'<h4>POTENTIAL VULNERABILITIES ('+str(len(host.get("vulns",[])))+')</h4><table><tr><th>Sev</th><th>Finding</th><th>Detail</th><th>CVE</th><th>Port</th></tr>'+vls+'</table>' if host.get('vulns') else '<p style="color:#00ff41">[ NO VULNERABILITIES DETECTED ]</p>'}
          {todos_html}
        </div>"""
    tc = sum(len([p for p in h.get('ports',[]) if p['classification']=='critical']) for h in scan.get('hosts',[]))
    tv = sum(len(h.get('vulns',[])) for h in scan.get('hosts',[]))
    tt = sum(len(h.get('todos',[])) for h in scan.get('hosts',[]))
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>NmapViz Report {ts}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Share+Tech+Mono&display=swap');
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#000;color:#c9d1d9;font-family:'Share Tech Mono',monospace;padding:40px;line-height:1.6}}
  h1{{color:#00ff41;border-bottom:1px solid #00ff41;padding-bottom:8px;margin-bottom:16px;font-size:22px}}
  h2{{color:#00ff41;margin:28px 0 10px;font-size:16px}}
  h3{{color:#58a6ff;margin:16px 0 4px;font-size:14px}}
  h4{{color:#888;font-size:12px;text-transform:uppercase;letter-spacing:1px;margin:12px 0 6px}}
  .meta{{color:#888;font-size:12px;margin-bottom:8px}}
  .hb{{background:#0d1117;border:1px solid #30363d;border-left:3px solid #58a6ff;border-radius:4px;padding:18px;margin-bottom:18px}}
  table{{width:100%;border-collapse:collapse;font-size:12px;margin:8px 0}}
  th{{background:#161b22;padding:7px 10px;text-align:left;color:#888;font-size:11px;text-transform:uppercase;letter-spacing:.5px}}
  td{{padding:6px 10px;border-bottom:1px solid #21262d;vertical-align:top}}
  .stats{{display:flex;gap:24px;background:#0d1117;border:1px solid #30363d;padding:16px;margin:16px 0;border-radius:4px;flex-wrap:wrap}}
  .stat{{text-align:center}}.val{{font-size:26px;font-weight:700;color:#00ff41}}.lbl{{font-size:11px;color:#888;text-transform:uppercase}}
  code{{color:#58a6ff;background:rgba(88,166,255,.1);padding:2px 6px;border-radius:3px}}
  footer{{margin-top:40px;color:#555;font-size:11px;border-top:1px solid #30363d;padding-top:16px}}
</style></head><body>
<h1>&gt;_ NMAPVIZ // SCAN REPORT</h1>
<p><strong>Date:</strong> {ts} &nbsp;|&nbsp; <strong>Source:</strong> {files}</p>
<p><strong>Command:</strong> <code>{scan.get('args','N/A')}</code></p>
<p><strong>Version detection:</strong> {'YES (-sV)' if scan.get('has_version') else 'NO (basic scan)'} &nbsp;|&nbsp;
   <strong>NSE Scripts:</strong> {'YES' if scan.get('has_scripts') else 'NO'}</p>
<div class='stats'>
  <div class='stat'><div class='val'>{len(scan.get('hosts',[]))}</div><div class='lbl'>Hosts</div></div>
  <div class='stat'><div class='val'>{sum(len(h.get('ports',[])) for h in scan.get('hosts',[]))}</div><div class='lbl'>Open Ports</div></div>
  <div class='stat'><div class='val' style='color:#ff3333'>{tc}</div><div class='lbl'>Critical Ports</div></div>
  <div class='stat'><div class='val' style='color:#bc8cff'>{tv}</div><div class='lbl'>Vulnerabilities</div></div>
  <div class='stat'><div class='val' style='color:#e07a30'>{tt}</div><div class='lbl'>Actions</div></div>
</div>
<h2>// HOST DETAILS</h2>{hosts_html}
<footer>Generated by NmapViz &middot; https://github.com/YOUR_USERNAME/nmap-visualizer</footer>
</body></html>"""


# ─────────────────────────────────────────────────────────────────────────────
# REPORT: MARKDOWN
# ─────────────────────────────────────────────────────────────────────────────

SEV_REFERENCE_MD = """
---

## Severity Classification Reference

| Level | Criteria | Examples |
|-------|----------|---------|
| **CRITICAL** | Confirmed CVEs with unauthenticated RCE / auth bypass, or cleartext legacy protocols that must never be exposed | EternalBlue, Heartbleed, Telnet, Cisco Smart Install, IPMI cipher-0 bypass |
| **HIGH** | Services or configurations that are dangerous when internet-exposed or reachable by untrusted users; may enable lateral movement | FTP anonymous, outdated OpenSSH, exposed RDP, SMB signing disabled |
| **MEDIUM** | Misconfigurations requiring deliberate exploitation or insider access; no direct unauthenticated path | RDP without NLA, SMBv2 signing not required, weak RSA keys |
| **LOW** | Best-practice deviations with limited direct impact; harden as time permits | SSH password auth enabled, expired TLS certificate |
| **INFO** | Observations with no direct exploitability; context for the auditor | Port-only service identification, version detection gaps |

> Severity ratings in this report reflect the inherent risk of the detected condition. They may require adjustment based on network segmentation, compensating controls, and business context. Ratings can be overridden in the NmapViz interface.
"""


def generate_markdown_report(scan, meta):
    ts    = meta.get('timestamp_display', '')
    files = ', '.join(meta.get('filenames', []))
    lines = [
        "# NmapViz Scan Report", "",
        f"**Generated:** {ts}  ", f"**Source files:** {files}  ",
        f"**Command:** `{scan.get('args','N/A')}`  ",
        f"**Version detection:** {'Yes (-sV)' if scan.get('has_version') else 'No (basic scan)'}  ", "",
        "## Summary", "",
        "| Metric | Value |", "|--------|-------|",
        f"| Active Hosts | {len(scan.get('hosts',[]))} |",
        f"| Open Ports | {sum(len(h.get('ports',[])) for h in scan.get('hosts',[]))} |",
        f"| Critical Ports | {sum(len([p for p in h.get('ports',[]) if p['classification']=='critical']) for h in scan.get('hosts',[]))} |",
        f"| Vulnerabilities | {sum(len(h.get('vulns',[])) for h in scan.get('hosts',[]))} |", "",
        "---", "", "## Hosts", "",
    ]
    sev_order = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFO':4}
    for host in scan.get('hosts', []):
        lines += [f"### {host.get('ip','')}  {host.get('hostname','')}", ""]
        if host.get('os'): lines.append(f"- **OS:** {host['os']} ({host.get('os_accuracy',0)}%)")
        if host.get('mac'): lines.append(f"- **MAC:** {host['mac']} {host.get('vendor','')}")
        lines += [f"- **Open Ports:** {len(host.get('ports',[]))}", ""]
        if host.get('ports'):
            lines += ["#### Open Ports", "", "| Port | Service | Version | Class |",
                      "|------|---------|---------|-------|"]
            for p in sorted(host['ports'], key=lambda x:{'critical':0,'interesting':1,'normal':2}.get(x['classification'],2)):
                svc = p.get('service',{})
                ver = ' '.join(filter(None,[svc.get('product',''),svc.get('version','')]))
                icon = {'critical':'🔴','interesting':'🟠','normal':'🟢'}.get(p['classification'],'')
                lines.append(f"| {p['portid']}/{p['protocol']} | {svc.get('name','')} | {ver} | {icon} {p['classification']} |")
            lines.append("")
        if host.get('vulns'):
            lines += ["#### ⚠️ Vulnerabilities", ""]
            for v in sorted(host['vulns'], key=lambda x:sev_order.get(x.get('severity','INFO'),9)):
                lines += [f"**[{v['severity']}] {v['title']}** (Port {v['port']})", "",
                          f"> {v['description']}", f"> CVE: `{v.get('cve','N/A')}`", ""]
        lines += ["---", ""]
    lines.append("*Generated by [NmapViz](https://github.com/YOUR_USERNAME/nmap-visualizer)*")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# REPORT: PDF (Executive Summary)
# ─────────────────────────────────────────────────────────────────────────────

def _pdf_safe(text):
    """Replace characters unsupported by Helvetica with ASCII equivalents."""
    if not text:
        return ''
    replacements = {
        '—': '--',  # em dash (U+2014)
        '–': '-',    # en dash
        '‘': "'",    # left single quote
        '’': "'",    # right single quote
        '“': '"',    # left double quote
        '”': '"',    # right double quote
        '•': '*',    # bullet
        '…': '...',  # ellipsis
        'é': 'e',    # e acute
        'à': 'a',    # a grave
        'ü': 'u',    # u umlaut
        'ö': 'o',    # o umlaut
        'ä': 'a',    # a umlaut
        'è': 'e',    # e grave
        'á': 'a',    # a acute
        'ó': 'o',    # o acute
        'ú': 'u',    # u acute
        'í': 'i',    # i acute
        '·': '.',      # middle dot
        '’': "'",
        '×': 'x',    # multiplication sign
        '°': 'deg',  # degree sign
        '→': '->',   # right arrow
        '←': '<-',   # left arrow
        '▶': '>',    # triangle
        '≈': '~',    # approximately
    }
    result = str(text)
    for char, replacement in replacements.items():
        result = result.replace(char, replacement)
    # Final pass: strip any remaining non-latin1 chars
    return result.encode('latin-1', errors='replace').decode('latin-1')



def generate_xlsx_report(scan, meta):
    """Generate a multi-sheet xlsx workbook with full scan data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ts    = meta.get('timestamp_display', '')
    files = ', '.join(meta.get('filenames', []))

    # ── Colour palette (light/professional) ────────────────────────────────
    C_BG_HEADER = 'FF2C3E50'   # dark header row
    C_BG_ROW1   = 'FFFFFFFF'   # white
    C_BG_ROW2   = 'FFF5F7FA'   # very light grey alternate
    C_TEXT_HDR  = 'FFFFFFFF'   # white text on dark header
    C_TEXT      = 'FF1A1A2E'   # near-black body text
    C_MUTED     = 'FF555B6E'   # grey secondary text
    # Severity accent colours (used for text only, not background)
    C_CRIT      = 'FF6B21A8'   # purple
    C_HIGH      = 'FFDC2626'   # red
    C_MED       = 'FFD97706'   # amber
    C_LOW       = 'FF16A34A'   # green
    C_INFO      = 'FF2563EB'   # blue
    C_ACCENT    = 'FF1D4ED8'   # blue accent

    sev_colors = {'CRITICAL': C_CRIT, 'HIGH': C_HIGH, 'MEDIUM': C_MED,
                  'LOW': C_LOW, 'INFO': C_INFO}

    def hdr_font(bold=True):
        return Font(name='Calibri', bold=bold, color=C_TEXT_HDR, size=11)
    def cell_font(color=C_TEXT, bold=False, size=10):
        return Font(name='Calibri', color=color, bold=bold, size=size)
    def hdr_fill():
        return PatternFill('solid', fgColor=C_BG_HEADER)
    def row_fill(i):
        return PatternFill('solid', fgColor=C_BG_ROW1 if i%2==0 else C_BG_ROW2)
    def set_col_width(ws, col, width):
        ws.column_dimensions[get_column_letter(col)].width = width

    def write_header(ws, headers, widths):
        ws.append(headers)
        for i, cell in enumerate(ws[1], 1):
            cell.font   = hdr_font()
            cell.fill   = hdr_fill()
            cell.alignment = Alignment(horizontal='left', vertical='center')
        for i, w in enumerate(widths, 1):
            set_col_width(ws, i, w)
        ws.row_dimensions[1].height = 20

    hosts = scan.get('hosts', [])

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = '4D9DE0'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    summary_data = [
        ('Report Generated', ts),
        ('Source Files', files),
        ('Scan Command', scan.get('args', 'N/A')),
        ('Version Detection', 'Yes (-sV)' if scan.get('has_version') else 'No (basic scan)'),
        ('NSE Scripts', 'Yes' if scan.get('has_scripts') else 'No'),
        ('', ''),
        ('Active Hosts', len(hosts)),
        ('Total Open Ports', sum(len(h.get('ports',[])) for h in hosts)),
        ('Critical Ports', sum(len([p for p in h.get('ports',[]) if p['classification']=='critical']) for h in hosts)),
        ('Interesting Ports', sum(len([p for p in h.get('ports',[]) if p['classification']=='interesting']) for h in hosts)),
        ('Total Findings', sum(len(h.get('vulns',[])) for h in hosts)),
        ('Critical Findings', sum(len([v for v in h.get('vulns',[]) if v['severity']=='CRITICAL']) for h in hosts)),
        ('High Findings', sum(len([v for v in h.get('vulns',[]) if v['severity']=='HIGH']) for h in hosts)),
        ('Investigation Actions', sum(len(h.get('todos',[])) for h in hosts)),
    ]
    for i, (k, v) in enumerate(summary_data, 1):
        ws.cell(i, 1, k).font  = cell_font(C_MUTED, bold=True)
        ws.cell(i, 2, str(v)).font = cell_font(C_TEXT)
        ws.cell(i, 1).fill = row_fill(i)
        ws.cell(i, 2).fill = row_fill(i)
        ws.row_dimensions[i].height = 16

    # ── Sheet 2: Hosts ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Hosts')
    ws2.sheet_properties.tabColor = '4D9DE0'
    cols2 = ['IP Address','Hostname','OS','OS Accuracy','MAC','Vendor','Open Ports','Critical Ports','Interesting Ports','Findings','Investigation Actions']
    wids2 = [18,28,40,14,20,20,14,16,18,12,22]
    write_header(ws2, cols2, wids2)
    sev_order = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFO':4}
    for i, h in enumerate(sorted(hosts, key=lambda x: [int(p) for p in (x.get('ip','0')+'.0.0.0').split('.')[:4]]), 2):
        hc = len([p for p in h.get('ports',[]) if p['classification']=='critical'])
        hi = len([p for p in h.get('ports',[]) if p['classification']=='interesting'])
        hv = len(h.get('vulns',[]))
        ht = len(h.get('todos',[]))
        row = [h.get('ip',''), h.get('hostname',''), h.get('os',''), h.get('os_accuracy',''),
               h.get('mac',''), h.get('vendor',''), len(h.get('ports',[])), hc, hi, hv, ht]
        ws2.append(row)
        f = row_fill(i)
        for j, cell in enumerate(ws2[i], 1):
            cell.fill = f
            cell.alignment = Alignment(vertical='center')
            if j==1: cell.font = cell_font(C_ACCENT, bold=True)
            elif j==8 and hc: cell.font = cell_font(C_CRIT, bold=True)
            elif j==9 and hi: cell.font = cell_font(C_HIGH)
            elif j==10 and hv: cell.font = cell_font(C_MED)
            else: cell.font = cell_font(C_MUTED if j>3 else C_TEXT)
        ws2.row_dimensions[i].height = 16

    # ── Sheet 3: Ports ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Ports')
    ws3.sheet_properties.tabColor = '3FB950'
    cols3 = ['Host IP','Port','Protocol','Classification','Service Name','Product','Version','Extra Info','Note']
    wids3 = [18,8,10,14,18,24,20,24,50]
    write_header(ws3, cols3, wids3)
    row_i = 2
    cls_order = {'critical':0,'interesting':1,'normal':2}
    all_ports = []
    for h in hosts:
        for p in h.get('ports',[]):
            all_ports.append((h.get('ip',''), p))
    all_ports.sort(key=lambda x: (cls_order.get(x[1]['classification'],2), x[0], x[1]['portid']))
    for ip, p in all_ports:
        svc = p.get('service',{})
        cls = p.get('classification','normal')
        row = [ip, p['portid'], p['protocol'], cls.upper(),
               svc.get('name',''), svc.get('product',''), svc.get('version',''), svc.get('extrainfo',''),
               p.get('class_reason','')]
        ws3.append(row)
        f = row_fill(row_i)
        for j, cell in enumerate(ws3[row_i], 1):
            cell.fill = f
            cell.alignment = Alignment(vertical='center', wrap_text=(j==9))
            if j==1: cell.font = cell_font(C_ACCENT, bold=True)
            elif j==4:
                col = {C_CRIT:'critical',C_HIGH:'interesting'}.get
                c = C_CRIT if cls=='critical' else C_HIGH if cls=='interesting' else C_LOW
                cell.font = cell_font(c, bold=True)
            else: cell.font = cell_font(C_MUTED if j>5 else C_TEXT)
        ws3.row_dimensions[row_i].height = 16
        row_i += 1

    # ── Sheet 4: Findings ────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Findings')
    ws4.sheet_properties.tabColor = C_CRIT[2:]
    cols4 = ['Host IP','Severity','Title','CVE','Port','Description','Source']
    wids4 = [18,12,40,20,8,60,22]
    write_header(ws4, cols4, wids4)
    row_i = 2
    all_vulns = []
    for h in hosts:
        for v in h.get('vulns',[]):
            all_vulns.append((h.get('ip',''), v))
    all_vulns.sort(key=lambda x: (sev_order.get(x[1].get('severity','INFO'),9), x[0]))
    for ip, v in all_vulns:
        row = [ip, v.get('severity',''), v.get('title',''), v.get('cve','N/A'),
               v.get('port',''), v.get('description',''), v.get('script','')]
        ws4.append(row)
        f = row_fill(row_i)
        sev_col = sev_colors.get(v.get('severity','INFO'), C_MUTED)
        for j, cell in enumerate(ws4[row_i], 1):
            cell.fill = f
            cell.alignment = Alignment(vertical='center', wrap_text=(j==6))
            if j==1: cell.font = cell_font(C_ACCENT, bold=True)
            elif j==2: cell.font = cell_font(sev_col, bold=True)
            elif j==3: cell.font = cell_font(C_TEXT, bold=True)
            else: cell.font = cell_font(C_MUTED)
        ws4.row_dimensions[row_i].height = 16 if len(v.get('description',''))<80 else 28
        row_i += 1

    # ── Sheet 5: Investigation Actions (TODOs) ───────────────────────────────
    ws5 = wb.create_sheet('Investigation Actions')
    ws5.sheet_properties.tabColor = C_MED[2:]
    cols5 = ['Host IP','Priority','Port','Title','Reason','Description','Follow-up Scan']
    wids5 = [18,10,8,38,40,55,70]
    write_header(ws5, cols5, wids5)
    row_i = 2
    for h in hosts:
        for t in h.get('todos',[]):
            row = [h.get('ip',''), t.get('priority','').upper(), t.get('port',''),
                   t.get('title',''), t.get('reason',''), t.get('description',''), t.get('followup','')]
            ws5.append(row)
            f = row_fill(row_i)
            pri_col = C_HIGH if t.get('priority')=='high' else C_MED
            for j, cell in enumerate(ws5[row_i], 1):
                cell.fill = f
                cell.alignment = Alignment(vertical='center', wrap_text=(j in (5,6,7)))
                if j==1: cell.font = cell_font(C_ACCENT, bold=True)
                elif j==2: cell.font = cell_font(pri_col, bold=True)
                elif j==7: cell.font = cell_font(C_INFO, size=9)
                else: cell.font = cell_font(C_TEXT if j<5 else C_MUTED)
            ws5.row_dimensions[row_i].height = 16
            row_i += 1

    buf = __import__('io').BytesIO()

    # ── Sheet 6: Severity Classification Reference ───────────────────────────
    ws6 = wb.create_sheet('Severity Reference')
    ws6.sheet_properties.tabColor = C_INFO[2:]
    ws6.column_dimensions['A'].width = 14
    ws6.column_dimensions['B'].width = 70
    ws6.column_dimensions['C'].width = 65

    ws6.append(['Level', 'Criteria', 'Examples'])
    for i, cell in enumerate(ws6[1], 1):
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[1].height = 20
    set_col_width(ws6, 1, 14); set_col_width(ws6, 2, 70); set_col_width(ws6, 3, 65)

    sev_ref_rows = [
        ('CRITICAL', C_CRIT,
         'Confirmed CVEs with unauthenticated RCE / auth bypass, or cleartext legacy protocols that must never be exposed.',
         'EternalBlue (MS17-010), Heartbleed (CVE-2014-0160), Telnet, IPMI cipher-0 bypass, Cisco Smart Install, Ghostcat'),
        ('HIGH', C_HIGH,
         'Services or configurations dangerous when internet-exposed or reachable by untrusted users. May enable lateral movement.',
         'FTP anonymous access, exposed RDP, outdated OpenSSH, SMB signing disabled, POODLE (CVE-2014-3566)'),
        ('MEDIUM', C_MED,
         'Misconfigurations requiring network position or deliberate attack chain. No direct unauthenticated exploitation path.',
         'RDP without NLA, SMBv2 signing not required, weak RSA host keys (1024-bit), expired TLS certificate'),
        ('LOW', C_LOW,
         'Best-practice deviations with limited direct impact. Address during routine hardening cycles.',
         'SSH password authentication enabled, open HTTP redirect'),
        ('INFO', C_INFO,
         'Informational observations with no direct exploitability. Useful context for scoping further tests.',
         'Port-only service identification (no -sV used), version detection gaps from basic scans'),
    ]
    for ri, (level, col, criteria, examples) in enumerate(sev_ref_rows, 2):
        ws6.append([level, criteria, examples])
        f = row_fill(ri)
        for j, cell in enumerate(ws6[ri], 1):
            cell.fill = f
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if j == 1: cell.font = cell_font(col, bold=True, size=11)
            elif j == 2: cell.font = cell_font(C_TEXT)
            else: cell.font = cell_font(C_MUTED)
        ws6.row_dimensions[ri].height = 50

    ws6.append([''])
    ws6.append(['Note', 'Severity ratings reflect inherent risk and may require adjustment based on network segmentation, compensating controls, and business context. Ratings can be overridden in the NmapViz interface before generating reports.'])
    note_row = ws6.max_row
    ws6.cell(note_row, 1).font = cell_font(C_MUTED, bold=True)
    ws6.cell(note_row, 2).font = cell_font(C_MUTED)
    ws6.cell(note_row, 2).alignment = Alignment(wrap_text=True)
    ws6.row_dimensions[note_row].height = 36

    wb.save(buf)
    return buf.getvalue()


def generate_pdf_report(scan, meta):
    """Professional A4 executive summary PDF - clean layout, no overflow."""
    from fpdf import FPDF, XPos, YPos

    ts    = _pdf_safe(meta.get('timestamp_display', 'N/A'))
    files = _pdf_safe(', '.join(meta.get('filenames', [])))
    hosts = scan.get('hosts', [])

    total_hosts = len(hosts)
    total_ports = sum(len(h.get('ports',[])) for h in hosts)
    crit_ports  = sum(len([p for p in h.get('ports',[]) if p['classification']=='critical']) for h in hosts)
    inter_ports = sum(len([p for p in h.get('ports',[]) if p['classification']=='interesting']) for h in hosts)
    total_vulns = sum(len(h.get('vulns',[])) for h in hosts)
    crit_vulns  = sum(len([v for v in h.get('vulns',[]) if v['severity']=='CRITICAL']) for h in hosts)
    high_vulns  = sum(len([v for v in h.get('vulns',[]) if v['severity']=='HIGH']) for h in hosts)
    med_vulns   = sum(len([v for v in h.get('vulns',[]) if v['severity']=='MEDIUM']) for h in hosts)
    low_vulns   = sum(len([v for v in h.get('vulns',[]) if v['severity']=='LOW']) for h in hosts)
    total_todos = sum(len(h.get('todos',[])) for h in hosts)

    risk_score = crit_vulns*10 + high_vulns*5 + med_vulns*2 + crit_ports*3
    risk_label = 'CRITICAL' if risk_score>=60 else 'HIGH' if risk_score>=20 else 'MEDIUM' if risk_score>0 else 'LOW'

    # ── Colour palette (clean, printable) ────────────────────────────────────
    W  = (255, 255, 255)      # white
    BG = (248, 250, 252)      # near-white bg
    BK = (25,  25,  40)       # near-black body text
    GR = (100, 110, 120)      # medium grey
    LG = (200, 210, 218)      # light grey line
    AC = (44,  105, 190)      # blue accent
    CRIT = (100, 40,  200)    # purple
    HIGH = (200, 40,  40)     # red
    MED  = (200, 120, 20)     # amber
    LOW  = (30,  150, 60)     # green
    INFO = (44,  105, 190)    # blue
    RISK_COL = {'LOW':LOW, 'MEDIUM':MED, 'HIGH':HIGH, 'CRITICAL':CRIT}
    SEV_COL  = {'CRITICAL':CRIT, 'HIGH':HIGH, 'MEDIUM':MED, 'LOW':LOW, 'INFO':INFO}

    MARGIN_L = 18
    MARGIN_R = 18
    PAGE_W   = 210
    USABLE_W = PAGE_W - MARGIN_L - MARGIN_R   # 174 mm

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=False)

    def new_page():
        pdf.add_page()
        pdf.set_margins(MARGIN_L, 15, MARGIN_R)
        pdf.set_fill_color(*W)
        pdf.rect(0, 0, 210, 297, 'F')

    def hline(y=None, col=LG):
        y = y or pdf.get_y()
        pdf.set_draw_color(*col)
        pdf.line(MARGIN_L, y, PAGE_W - MARGIN_R, y)

    def section_header(title, y=None):
        if y is None:
            pdf.ln(5)
        else:
            pdf.set_y(y)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(*AC)
        pdf.set_fill_color(*BG)
        pdf.cell(USABLE_W, 7, '  ' + title.upper(), fill=True, new_x=XPos.LEFT, new_y=YPos.NEXT)
        hline(col=AC)
        pdf.ln(2)
        pdf.set_text_color(*BK)

    def page_footer():
        pdf.set_y(-13)
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(*GR)
        pdf.cell(0, 5, f'NmapViz  |  {ts}  |  Page {pdf.page_no()}  |  Confidential', align='C')

    def truncate(text, max_chars):
        t = _pdf_safe(str(text or ''))
        return t[:max_chars] + ('...' if len(t) > max_chars else '')

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE 1: COVER / EXECUTIVE OVERVIEW
    # ─────────────────────────────────────────────────────────────────────────
    new_page()

    # Top accent bar 6mm
    pdf.set_fill_color(*AC)
    pdf.rect(0, 0, 210, 6, 'F')

    # ── Header: Title left, Risk badge right ─────────────────────────────────
    pdf.set_xy(MARGIN_L, 12)
    pdf.set_font('Helvetica', 'B', 22)
    pdf.set_text_color(*AC)
    pdf.cell(110, 10, 'NmapViz')
    # Risk badge (right-aligned, 50mm wide)
    rc = RISK_COL[risk_label]
    pdf.set_fill_color(*rc)
    pdf.set_text_color(*W)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_xy(PAGE_W - MARGIN_R - 50, 12)
    pdf.cell(50, 10, f'RISK: {risk_label}', align='C', fill=True)

    pdf.set_xy(MARGIN_L, 23)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(*GR)
    pdf.cell(0, 5, 'Network Security Scan  |  Executive Summary Report')
    pdf.ln(2)
    hline(col=LG)
    pdf.ln(3)

    # ── Metadata: label left, value wraps if needed ───────────────────────────
    meta_items = [
        ('Date',               _pdf_safe(ts)),
        ('Files',              _pdf_safe(', '.join(meta.get('filenames', [])))),
        ('Version Detection',  'Yes (-sV)' if scan.get('has_version') else 'No (basic scan)'),
        ('NSE Scripts',        'Yes' if scan.get('has_scripts') else 'No'),
        ('Command',            _pdf_safe(scan.get('args', 'N/A'))),
    ]
    LABEL_W  = 32          # fixed label column
    VALUE_W  = USABLE_W - LABEL_W  # remaining width for value (wraps)
    LINE_H   = 4.5

    for label, value in meta_items:
        y_before = pdf.get_y()
        # Draw label (never wraps)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(*GR)
        pdf.set_xy(MARGIN_L, y_before)
        pdf.cell(LABEL_W, LINE_H, label + ':')
        # Draw value with multi_cell so it wraps within the margin
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*BK)
        pdf.set_xy(MARGIN_L + LABEL_W, y_before)
        pdf.multi_cell(VALUE_W, LINE_H, value, new_x=XPos.LEFT, new_y=YPos.NEXT)
        # If multi_cell advanced Y more than one line, label needs no adjustment
        # (it already rendered at y_before)
    pdf.ln(3)
    hline(col=LG)
    pdf.ln(4)

    # ── Stats grid: 3 columns x 2 rows, fixed absolute positions ─────────────
    section_header('SCAN OVERVIEW')
    stats = [
        ('Active Hosts',          str(total_hosts),  AC),
        ('Open Ports',            str(total_ports),  AC),
        ('Critical Ports',        str(crit_ports),   CRIT),
        ('Interesting Ports',     str(inter_ports),  MED),
        ('Total Findings',        str(total_vulns),  HIGH),
        ('Investigation Actions', str(total_todos),  MED),
    ]
    CELL_W = USABLE_W / 3 - 2   # ~56 mm each
    CELL_H = 18
    base_y  = pdf.get_y()
    for i, (label, val, col) in enumerate(stats):
        col_idx = i % 3
        row_idx = i // 3
        x = MARGIN_L + col_idx * (CELL_W + 2)
        y = base_y + row_idx * (CELL_H + 2)
        pdf.set_fill_color(*BG)
        pdf.rect(x, y, CELL_W, CELL_H, 'F')
        pdf.set_text_color(*col)
        pdf.set_font('Helvetica', 'B', 20)
        pdf.set_xy(x + 3, y + 2)
        pdf.cell(CELL_W - 6, 10, val)
        pdf.set_text_color(*GR)
        pdf.set_font('Helvetica', '', 7)
        pdf.set_xy(x + 3, y + 12)
        pdf.cell(CELL_W - 6, 5, label.upper())
    # Advance past the 2-row grid
    pdf.set_y(base_y + 2 * (CELL_H + 2) + 4)

    # ── Severity bar chart ────────────────────────────────────────────────────
    hline(col=LG); pdf.ln(4)
    section_header('FINDINGS BY SEVERITY')
    sev_rows = [
        ('CRITICAL', crit_vulns, CRIT),
        ('HIGH',     high_vulns, HIGH),
        ('MEDIUM',   med_vulns,  MED),
        ('LOW',      low_vulns,  LOW),
    ]
    bar_max  = max(1, total_vulns)
    BAR_AREA = USABLE_W - 50   # reserve 50mm for label+count
    for sev, count, col in sev_rows:
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(*col)
        pdf.cell(22, 5.5, sev)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*BK)
        pdf.cell(10, 5.5, str(count), align='R')
        pdf.cell(4, 5.5, '')   # spacer
        bar_len = int((count / bar_max) * BAR_AREA) if count else 0
        if bar_len > 0:
            pdf.set_fill_color(*col)
            pdf.rect(pdf.get_x(), pdf.get_y() + 1, bar_len, 3.5, 'F')
        pdf.ln(6)

    page_footer()

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE 2: TOP FINDINGS (CRITICAL + HIGH)
    # ─────────────────────────────────────────────────────────────────────────
    new_page()
    section_header('CRITICAL, HIGH AND MEDIUM SEVERITY FINDINGS')

    all_vulns_chm = []
    for h in hosts:
        for v in h.get('vulns', []):
            if v['severity'] in ('CRITICAL', 'HIGH', 'MEDIUM'):
                all_vulns_chm.append({**v, 'host_ip': h['ip']})
    sev_order = {'CRITICAL':0,'HIGH':1,'MEDIUM':2,'LOW':3,'INFO':4}
    all_vulns_chm.sort(key=lambda x: sev_order.get(x['severity'], 9))

    if not all_vulns_chm:
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(*LOW)
        pdf.cell(0, 8, 'No Critical, High or Medium severity findings detected in this scan.')
    else:
        for v in all_vulns_chm[:30]:
            if pdf.get_y() > 262:
                page_footer()
                new_page()
                section_header('FINDINGS (continued)')
            sc   = SEV_COL.get(v['severity'], GR)
            y0   = pdf.get_y()
            ITEM_H = 14   # fixed item height
            # Left colour strip
            pdf.set_fill_color(*sc)
            pdf.rect(MARGIN_L, y0, 2.5, ITEM_H, 'F')
            # Severity label
            pdf.set_xy(MARGIN_L + 4, y0 + 0.5)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*sc)
            pdf.cell(20, 4.5, v['severity'])
            # Host + port right-aligned
            meta_str = f"{v['host_ip']}  port {v['port']}"
            if v.get('cve') and v['cve'] != 'N/A':
                meta_str += f"  {v['cve']}"
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*GR)
            pdf.cell(0, 4.5, meta_str, align='R')
            # Title on next line
            pdf.set_xy(MARGIN_L + 4, y0 + 5.5)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*BK)
            pdf.cell(0, 4.5, truncate(v['title'], 90))
            # Description truncated to one line
            pdf.set_xy(MARGIN_L + 4, y0 + 10)
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*GR)
            pdf.cell(0, 4, truncate(v['description'], 120))
            pdf.set_y(y0 + ITEM_H + 1)
            hline(col=(235, 238, 242))

    page_footer()

    # ─────────────────────────────────────────────────────────────────────────
    # PAGE 3+: HOST SUMMARY TABLE
    # ─────────────────────────────────────────────────────────────────────────
    new_page()
    section_header('HOST SUMMARY')

    # Table column widths (must sum to USABLE_W = 174)
    COL_WIDTHS  = [32, 46, 14, 16, 18, 14, 16, 18]
    COL_HEADERS = ['IP Address', 'OS / Platform', 'Ports', 'Critical', 'Interesting', 'Vulns', 'Actions', 'Top Sev']

    def draw_table_header(ws):
        pdf.set_fill_color(44, 62, 80)
        for j, (h, w) in enumerate(zip(COL_HEADERS, COL_WIDTHS)):
            pdf.set_xy(MARGIN_L + sum(COL_WIDTHS[:j]), pdf.get_y())
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_text_color(*W)
            pdf.cell(w, 5.5, h, fill=True)
        pdf.ln(6)

    draw_table_header(None)
    sorted_hosts = sorted(hosts, key=lambda h: [int(x) for x in (h.get('ip','0')+'.0.0.0').split('.')[:4]])

    for row_i, host in enumerate(sorted_hosts):
        if pdf.get_y() > 262:
            page_footer()
            new_page()
            section_header('HOST SUMMARY (continued)')
            draw_table_header(None)

        hc = len([p for p in host.get('ports',[]) if p['classification']=='critical'])
        hi = len([p for p in host.get('ports',[]) if p['classification']=='interesting'])
        hv = len(host.get('vulns',[]))
        ht = len(host.get('todos',[]))
        # Top severity of this host
        host_sevs = [v.get('severity','INFO') for v in host.get('vulns',[])]
        top_sev   = min(host_sevs, key=lambda s: sev_order.get(s,9)) if host_sevs else 'clean'
        top_col   = SEV_COL.get(top_sev, LOW) if top_sev != 'clean' else LOW

        fill_bg = BG if row_i % 2 == 0 else W
        pdf.set_fill_color(*fill_bg)

        values = [
            _pdf_safe(host.get('ip','?')),
            truncate(host.get('os','Unknown') or 'Unknown', 30),
            str(len(host.get('ports',[]))),
            str(hc) if hc else '-',
            str(hi) if hi else '-',
            str(hv) if hv else '-',
            str(ht) if ht else '-',
            top_sev,
        ]
        text_colors = [
            AC, GR, BK,
            CRIT if hc else GR,
            MED  if hi else GR,
            HIGH if hv else GR,
            MED  if ht else GR,
            top_col,
        ]
        bolds = [True, False, False, bool(hc), bool(hi), bool(hv), bool(ht), bool(host_sevs)]

        for j, (val, w) in enumerate(zip(values, COL_WIDTHS)):
            pdf.set_xy(MARGIN_L + sum(COL_WIDTHS[:j]), pdf.get_y())
            pdf.set_font('Helvetica', 'B' if bolds[j] else '', 7)
            pdf.set_text_color(*text_colors[j])
            pdf.cell(w, 5, val, fill=True)
        pdf.ln(5.5)

    # ─────────────────────────────────────────────────────────────────────────
    # LAST PAGE: SEVERITY CLASSIFICATION REFERENCE
    # ─────────────────────────────────────────────────────────────────────────
    page_footer()
    new_page()
    section_header('SEVERITY CLASSIFICATION REFERENCE')
    pdf.ln(2)

    sev_ref = [
        ('CRITICAL', CRIT,
         'Confirmed CVEs with unauthenticated RCE / authentication bypass, or cleartext '
         'legacy protocols that must never be exposed under any circumstances.',
         'EternalBlue (MS17-010), Heartbleed, Telnet, IPMI cipher-0 bypass, Cisco Smart Install, Ghostcat'),
        ('HIGH', HIGH,
         'Services or configurations dangerous when internet-exposed or reachable by '
         'untrusted users. May enable lateral movement without deep exploitation.',
         'FTP anonymous access, exposed RDP, outdated OpenSSH, SMB signing disabled, POODLE'),
        ('MEDIUM', MED,
         'Misconfigurations requiring network position or deliberate attack chain. '
         'No direct unauthenticated exploitation path.',
         'RDP without NLA, SMBv2 signing not required, weak RSA host keys, expired TLS cert'),
        ('LOW', LOW,
         'Best-practice deviations with limited direct impact. Should be addressed '
         'during routine hardening cycles.',
         'SSH password authentication enabled, open HTTP redirect'),
        ('INFO', INFO,
         'Informational observations with no direct exploitability. Useful context '
         'for the auditor and for scoping further tests.',
         'Port-only service identification, version detection gaps from basic scans'),
    ]

    for sev, col, description, examples in sev_ref:
        y0 = pdf.get_y()
        # Coloured left bar
        pdf.set_fill_color(*col)
        pdf.rect(MARGIN_L, y0, 2.5, 20, 'F')
        # Severity label
        pdf.set_xy(MARGIN_L + 4, y0 + 1)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(*col)
        pdf.cell(0, 5, sev)
        # Description
        pdf.set_xy(MARGIN_L + 4, y0 + 7)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*BK)
        pdf.multi_cell(USABLE_W - 4, 4.2, _pdf_safe(description))
        # Examples
        pdf.set_x(MARGIN_L + 4)
        pdf.set_font('Helvetica', 'I', 7)
        pdf.set_text_color(*GR)
        pdf.cell(0, 4, 'e.g.: ' + examples)
        pdf.ln(4)
        hline(col=(235, 238, 242))
        pdf.ln(2)

    pdf.ln(4)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.set_text_color(*GR)
    pdf.multi_cell(0, 4.5, _pdf_safe(
        'Note: Severity ratings reflect the inherent risk of the detected condition and may require '
        'adjustment based on network segmentation, compensating controls, and business context. '
        'Ratings can be overridden manually in the NmapViz interface before generating reports.'))

    page_footer()
    return bytes(pdf.output())




# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    return jsonify({"status": "ok"})

@app.route('/api/parse', methods=['POST'])
def parse_files():
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({"error": "No files provided"}), 400
    scans, filenames = [], []
    for file in files:
        if not file.filename.lower().endswith('.xml'):
            return jsonify({"error": f"'{file.filename}' is not an XML file"}), 400
        try:
            content = file.read().decode('utf-8', errors='replace')
            scans.append(parse_nmap_xml(content))
            filenames.append(file.filename)
        except ET.ParseError as e:
            return jsonify({"error": f"Invalid XML in '{file.filename}': {str(e)}"}), 400
        except Exception as e:
            return jsonify({"error": f"Error processing '{file.filename}': {str(e)}"}), 500
    data = scans[0] if len(scans) == 1 else merge_scans(scans)
    meta = save_to_history(data, filenames)
    return jsonify({"success": True, "data": data, "scan_id": meta['id'], "merged": len(scans) > 1})

@app.route('/api/history')
def get_history():
    items = []
    for f in sorted(HISTORY_DIR.glob('*.json'), reverse=True):
        try:
            with open(f) as fp: d = json.load(fp)
            items.append({k: d[k] for k in
                ['id','timestamp','timestamp_display','filenames','host_count',
                 'total_ports','vuln_count','critical_ports','args'] if k in d})
        except Exception: continue
    return jsonify(items)

@app.route('/api/history/<scan_id>')
def get_scan(scan_id):
    path = HISTORY_DIR / f"{scan_id}.json"
    if not path.exists(): return jsonify({"error": "Scan not found"}), 404
    with open(path) as f: d = json.load(f)
    return jsonify({"success": True, "data": d['scan'],
                    "meta": {k: d[k] for k in ['id','timestamp_display','filenames','host_count','total_ports','vuln_count','critical_ports'] if k in d}})

@app.route('/api/history/<scan_id>', methods=['DELETE'])
def delete_scan(scan_id):
    path = HISTORY_DIR / f"{scan_id}.json"
    if path.exists(): path.unlink()
    return jsonify({"success": True})

@app.route('/api/report/<scan_id>/<fmt>')
def download_report(scan_id, fmt):
    path = HISTORY_DIR / f"{scan_id}.json"
    if not path.exists(): return jsonify({"error": "Scan not found"}), 404
    with open(path) as f: stored = json.load(f)
    scan = stored['scan']
    meta = {k: stored[k] for k in ['id','timestamp_display','filenames'] if k in stored}
    if fmt == 'json':
        return Response(json.dumps(scan, indent=2), mimetype='application/json',
            headers={"Content-Disposition": f"attachment; filename=nmapviz_{scan_id}.json"})
    elif fmt == 'html':
        return Response(generate_html_report(scan, meta), mimetype='text/html',
            headers={"Content-Disposition": f"attachment; filename=nmapviz_{scan_id}.html"})
    elif fmt == 'markdown':
        return Response(generate_markdown_report(scan, meta), mimetype='text/markdown',
            headers={"Content-Disposition": f"attachment; filename=nmapviz_{scan_id}.md"})
    elif fmt == 'pdf':
        try:
            pdf_bytes = generate_pdf_report(scan, meta)
            return Response(pdf_bytes, mimetype='application/pdf',
                headers={"Content-Disposition": f"attachment; filename=nmapviz_{scan_id}.pdf"})
        except Exception as e:
            return jsonify({"error": f"PDF generation failed: {str(e)}"}), 500
    elif fmt == 'xlsx':
        try:
            xlsx_bytes = generate_xlsx_report(scan, meta)
            return Response(xlsx_bytes, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={"Content-Disposition": f"attachment; filename=nmapviz_{scan_id}.xlsx"})
        except Exception as e:
            return jsonify({"error": f"XLSX generation failed: {str(e)}"}), 500
    else:
        return jsonify({"error": "Unknown format. Use: json, html, markdown, pdf, xlsx"}), 400

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=12221, debug=False)